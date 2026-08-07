from __future__ import annotations

import hashlib
import math
import tempfile
import unicodedata
from collections import defaultdict
from datetime import date, datetime
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Tuple

from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel


SOURCE_SHEET = "Tablas de datos 2026"
REQUIRED_COLUMNS = ("folio", "serie")


def normalize_text(value) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    # Repair the common UTF-8-as-Latin-1 artifact found in the source workbook.
    if "Ã" in text or "Â" in text:
        try:
            text = text.encode("latin-1").decode("utf-8")
        except UnicodeError:
            pass
    return " ".join(text.split())


def as_float(value) -> Optional[float]:
    if value is None or value == "":
        return None
    try:
        number = float(str(value).replace(",", "."))
        return number if math.isfinite(number) else None
    except (TypeError, ValueError):
        return None


def as_datetime(value) -> Optional[datetime]:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value
    if isinstance(value, date):
        return datetime.combine(value, datetime.min.time())
    number = as_float(value)
    if number is not None and number > 0:
        try:
            converted = from_excel(number)
            if isinstance(converted, datetime):
                return converted
            if isinstance(converted, date):
                return datetime.combine(converted, datetime.min.time())
        except (TypeError, ValueError, OverflowError):
            pass
    return None


def calculate_month(value) -> Tuple[Optional[int], str]:
    if value is None:
        return None, ""
    month_number = as_float(value)
    if month_number is not None and 1 <= int(month_number) <= 12:
        number = int(month_number)
        return number, str(number)
    text = normalize_text(value)
    names = {
        "enero": 1, "ene": 1, "febrero": 2, "feb": 2,
        "marzo": 3, "mar": 3, "abril": 4, "abr": 4,
        "mayo": 5, "may": 5, "junio": 6, "jun": 6,
        "julio": 7, "jul": 7, "agosto": 8, "ago": 8,
        "septiembre": 9, "sept": 9, "sep": 9, "octubre": 10,
        "oct": 10, "noviembre": 11, "nov": 11, "diciembre": 12,
        "dic": 12,
    }
    number = names.get(text.lower())
    return number, str(number) if number else text


def normalize_row(values: Dict[str, object]) -> Optional[Dict[str, object]]:
    folio = normalize_text(values.get("folio"))
    serie = normalize_text(values.get("serie"))
    if not folio or not serie:
        return None

    reception = as_datetime(values.get("fecha_recepcion"))
    cutting = as_datetime(values.get("fecha_corta"))
    age = as_float(values.get("Antigüedad"))
    if age is None and reception and cutting:
        age = (reception - cutting).total_seconds() / 86400

    month_number, month = calculate_month(values.get("Mes"))
    if month_number is None and reception:
        month_number, month = reception.month, str(reception.month)

    age_bucket = normalize_text(values.get("Estado"))
    if not age_bucket and age is not None:
        age_bucket = "<10 días" if age < 10 else ">10 días"

    return {
        "business_key": f"{folio}::{serie}",
        "folio": folio,
        "serie": serie,
        "secuencia": normalize_text(values.get("secuencia")) or None,
        "cod_producto": normalize_text(values.get("cod_producto")) or None,
        "fecha_recepcion": reception,
        "fecha_corta": cutting,
        "year": int(as_float(values.get("year")) or (reception.year if reception else 2026)),
        "month": month,
        "month_number": month_number,
        "product_length": as_float(values.get("largo_producto")),
        "age_days": age,
        "age_bucket": age_bucket or None,
        "wood_state": normalize_text(values.get("glosa_estado_madera")) or None,
        "zone": normalize_text(values.get("glosa_zona")) or None,
        "origin": normalize_text(values.get("glosa_origen")) or None,
        "destination": normalize_text(values.get("glosa_destino")) or None,
        "weight": as_float(values.get("Peso")),
    }


def read_source(path: Path) -> Iterable[Dict[str, object]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        if SOURCE_SHEET not in workbook.sheetnames:
            raise ValueError(f"No se encontró la hoja '{SOURCE_SHEET}'.")
        worksheet = workbook[SOURCE_SHEET]
        rows = worksheet.iter_rows(values_only=True)
        headers = [normalize_text(value) for value in next(rows)]
        header_map = {header: index for index, header in enumerate(headers) if header}
        missing = [column for column in REQUIRED_COLUMNS if column not in header_map]
        if missing:
            raise ValueError(f"Faltan columnas obligatorias: {', '.join(missing)}")

        aliases = {
            "fecha_recepcion": "fecha_recepcion",
            "fecha_corta": "fecha_corta",
            "year": "[Año]",
            "Mes": "Mes",
            "largo_producto": "largo_producto",
            "Antigüedad": "Antigüedad",
            "Estado": "Estado",
            "glosa_estado_madera": "glosa_estado_madera",
            "glosa_zona": "glosa_zona",
            "glosa_origen": "glosa_origen",
            "glosa_destino": "glosa_destino",
            "Peso": "Peso",
            "secuencia": "secuencia",
            "cod_producto": "cod_producto",
        }
        for row in rows:
            values = {
                target: row[header_map[source]] if source in header_map and header_map[source] < len(row) else None
                for target, source in aliases.items()
            }
            values["folio"] = row[header_map["folio"]] if header_map["folio"] < len(row) else None
            values["serie"] = row[header_map["serie"]] if header_map["serie"] < len(row) else None
            yield values
    finally:
        workbook.close()


def file_hash(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as source:
        for chunk in iter(lambda: source.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def save_upload(upload_file, suffix: str = ".xlsx") -> Tuple[Path, str]:
    with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as destination:
        while True:
            chunk = upload_file.file.read(1024 * 1024)
            if not chunk:
                break
            destination.write(chunk)
        path = Path(destination.name)
    return path, file_hash(path)
